import random
from time import process_time
from time import perf_counter
import timeit
from openpyxl import load_workbook




def merge(arr, n, mid, m):
    overall_num_comparisons = 0

    if m - n <= 0:
        return

    right_half_index = mid + 1
    left_half_index = n

    while left_half_index <= mid and right_half_index <= m:
        overall_num_comparisons += 1
        if arr[left_half_index] < arr[right_half_index]:
            left_half_index += 1

        elif arr[right_half_index] < arr[left_half_index]:
            temp = arr[right_half_index]
            for i in range(mid, left_half_index - 1, -1):
                arr[i + 1] = arr[i]
            arr[left_half_index] = temp
            left_half_index += 1
            right_half_index += 1
            mid += 1

        else:
            # if last elements then break
            if left_half_index == mid and right_half_index == m:
                break

            #
            left_half_index += 1

            temp = arr[right_half_index]
            for i in range(mid, left_half_index, -1):
                arr[i + 1] = arr[i]
            arr[left_half_index] = temp
            left_half_index += 1
            right_half_index += 1
            mid += 1
    return overall_num_comparisons


def mergeSort(arr, n, m):
    mid = (n + m) // 2
    if m - n <= 0:
        return
    else:
        mergeSort(arr, n, mid)
        mergeSort(arr, mid + 1, m)
    merge(arr, n, mid, m)

# # Simple test to check if merge sort is working
# arr = [i for i in range(20)]
# print(arr)
# random.shuffle(arr)
# print(arr)
# mergeSort(arr, 0, len(arr) - 1)
# print(arr)


def swap(index1, index2, arr):
    temp = arr[index1]
    arr[index1] = arr[index2]
    arr[index2] = temp


def basic_insertionSort(arr):
    num_of_comparisons = 0
    for i in range(1, len(arr)):
        for j in range(i, 0, -1):
            # print(i, j)
            num_of_comparisons += 1
            if arr[j] < arr[j - 1]:
                # print(arr, num_of_comparisons)
                swap(j, j - 1, arr)
                # print(arr)
            else:
                break

    return num_of_comparisons


def get_insertion_sort_time(arr):
    arr_length = len(arr)
    start = process_time()
    num_comparisons = basic_insertionSort(arr)
    # print(arr)
    stop = process_time()
    time_taken = stop - start
    print("Elapsed time for insertion sort in seconds:", time_taken)
    return num_comparisons

#
# # Simple test to check if insertion sort is working
# arr = [i for i in range(20)]
# print(arr)
# random.shuffle(arr)
# print(arr)
# num_comparisons = basic_insertionSort(arr)
# print(arr, num_comparisons)


def insertionSort(arr, first, last):
    num_of_comparisons = 0
    for i in range(first + 1, last + 1):
        for j in range(i, first, -1):
            # print(i, j)
            num_of_comparisons += 1
            if arr[j] < arr[j - 1]:
                # print(arr, num_of_comparisons)
                swap(j, j - 1, arr)
                # print(arr)
            else:
                break

    return num_of_comparisons


# arr = [i for i in range(20)]
# print(arr)
# random.shuffle(arr)
# print(arr)
# num_comparisons = insertionSort(arr, 10, len(arr) - 1)
# print(arr, num_comparisons)

def merge_insertion_sort(arr, first, last, switch_sort_num):
    num_comparison = 0
    # + 1 in both if and else because there is one comparison here in the if statement below
    if last - first > switch_sort_num:
        mid = (first + last) // 2
        num_comparison += merge_insertion_sort(arr, first, mid, switch_sort_num)
        num_comparison += merge_insertion_sort(arr, mid + 1, last, switch_sort_num)
        num_comparison += merge(arr, first, mid, last)
        return num_comparison

    else:
        num_comparison += insertionSort(arr, first, last)
        return num_comparison


# def merge_insertion_sort(arr, first, last, switch_sort_num):
#
#     # + 1 in both if and else because there is one comparison here in the if statement below
#     if last - first > switch_sort_num:
#         mid = (first + last) // 2
#         num_comparison = merge_insertion_sort(arr, first, mid, switch_sort_num)
#         num_comparison += merge_insertion_sort(arr, mid + 1, last, switch_sort_num)
#         return num_comparison + merge(arr, first, mid, last) + 1
#
#     else:
#         return 1 + insertionSort(arr, first, last)


def get_merge_insertion_sort_time(arr, switch_sort_num):
    arr_length = len(arr)
    # arr2 = arr.copy()
    # print(arr)
    #

    # start = perf_counter()
    # merge_insertion_sort(arr, 0, arr_length - 1, switch_sort_num)
    # stop = perf_counter()
    # #
    # time_taken = stop - start
    # print(arr)
    num_comparisons = merge_insertion_sort(arr.copy(), 0, arr_length - 1, switch_sort_num)
    repeat_loop_number = 5
    time_taken = min(timeit.repeat(lambda: merge_insertion_sort(arr.copy(), 0, arr_length - 1, switch_sort_num), repeat=5, number=repeat_loop_number)) / repeat_loop_number

    # print(arr)

    print("Switch sort num :", + switch_sort_num)
    print("Elapsed time during the whole program in seconds:", time_taken)
    print("Num of Comparisons :", num_comparisons)
    return time_taken, num_comparisons


def write_to_excel(row, col, data_size, S_values, comparison_arr, time_arr):
    wb = load_workbook("PythonOutput.xlsx")
    sheet = wb.active

    sheet.cell(row=row, column=col).value = "DataSize"
    sheet.cell(row=row, column=col + 1).value = data_size

    sheet.cell(row=row + 1, column=col).value = "S"
    for s_index in range(len(S_values)):
        sheet.cell(row=row + 1, column=col + s_index + 1).value = S_values[s_index]

    sheet.cell(row=row + 2, column=col).value = "Comparisons"
    for s_index in range(len(S_values)):
        sheet.cell(row=row + 2, column=col + s_index + 1).value = comparison_arr[s_index]

    sheet.cell(row=row + 3, column=col).value = "Time"
    for s_index in range(len(S_values)):
        sheet.cell(row=row + 3, column=col + s_index + 1).value = time_arr[s_index]

    wb.save("PythonOutput.xlsx")


def compare_diff_switch_sort_num(starting_num, ending_num, row, col):
    arr = [i for i in range(starting_num, ending_num + 1)]
    comparison_arr = []
    time_arr = []
    s_values = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27,
                28, 29, 30, 35, 40, 45, 50, 100, 200, 500, 1000]
    # print(arr)
    random.shuffle(arr)

    for S in s_values:
        temp_arr = arr.copy()
        # print("Before : ", temp_arr)
        time_taken, num_comparisons = get_merge_insertion_sort_time(temp_arr, S)
        time_arr.append(time_taken)
        comparison_arr.append(num_comparisons)
        # print("After : ", temp_arr)

    write_to_excel(row, col, ending_num - starting_num + 1, s_values, comparison_arr, time_arr)


# Compares S values and writes to excel for different data set size and S values
def main():
    row = 1
    col = 1
    data_size_arr = [10000]
    for data_size in data_size_arr:
        compare_diff_switch_sort_num(1, data_size, row, col)
        row += 10

main()


# arr = [x for x in range(1, 1001)]
# random.shuffle(arr)
# print(arr)
# num_comparisons = merge_insertion_sort(arr, 0, 1000 - 1, 0)
# print(arr)
# print(num_comparisons)
